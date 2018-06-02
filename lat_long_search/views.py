from django.http import HttpResponse
from django.template import loader
from openpyxl import load_workbook
from .forms import UploadFileForm
from django.shortcuts import render
from wsgiref.util import FileWrapper
from django.conf import settings
import os

# Create your views here.
def index(request):
	template = loader.get_template("lat_long_search/index.html")
	return HttpResponse(template.render({}, request))
	# return HttpResponse("Hello, world. You're at the polls index.")

def get_excel(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST, request.FILES)
		if form.is_valid():
			file_name = request.FILES["excel_file"].name
			if not is_accepted_extension(file_name):
				output1 = "Incorrect File extension"
			output1 = read_excel_file(request.FILES['excel_file'])
			return output1         
		else:
			output1 = "Form invalid"
	return HttpResponse(output1)

def is_accepted_extension(file_name):
	extension = file_name.split(".")[-1]
	if extension == "xlsx" or extension == "xls":
		return True
	return False

def read_excel_file(excel_file):
	# print(os.path.abspath(os.curdir))
	# local_excel_file_path = 'lat_long_search/media/sheets.xlsx'
	# with open(local_excel_file_path, 'wb+') as destination:
	# 	for chunk in excel_file.chunks():
	# 		destination.write(chunk)
	wb = load_workbook(excel_file)
	sheet1 = wb.worksheets[0]
	total_rows = sheet1.max_row
	list1 = []
	for i in range(1, total_rows+1):
		list1.append(sheet1.cell(column=1, row=i).value)
	print(list1)
	list1 = modify_list(list1)
	print(list1)
	# excel_file = local_excel_file_path
	print("New excel_file : ", excel_file)
	excel_file = write_excel_file(list1, excel_file)
	# print(excel_file.name)
	response = download_excel(excel_file)
	return response

def modify_list(list1):
	return list(map(lambda x: str(x) + " - 1", list1))

def write_excel_file(list1, excel_file):
	wb = load_workbook(excel_file)
	sheet1 = wb.worksheets[0]
	for i in range(len(list1)):
		print(list1[i])
		sheet1.cell(column=2, row=i+1).value  = list1[i]
	wb.save(excel_file)
	wb.close()
	return excel_file

def download_excel(excel_file):
	response = HttpResponse(FileWrapper(excel_file), content_type="application/vnd.ms-excel")
	response['Content-Disposition'] = 'inline; filename=' + excel_file.name
	return response